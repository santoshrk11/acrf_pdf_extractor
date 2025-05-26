"""Utility script to extract annotation data from an aCRF PDF.

Author: Santhosh RK
"""

import fitz  # PyMuPDF
import json
import os
import logging
import re
from dateutil import parser as date_parser  # For robust date parsing
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configure logging to write alongside this script using the script name
log_file = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    f"{os.path.splitext(os.path.basename(__file__))[0]}.log",
)
logging.basicConfig(
    level=logging.INFO,  # Set to logging.DEBUG for verbose output
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler(log_file), logging.StreamHandler()],
)


def parse_pdf_date(pdf_date_str):
    """Parse a PDF date string into a datetime object."""
    if not pdf_date_str:
        return None

    pdf_date_str = pdf_date_str.strip()
    if pdf_date_str in {"00000000000000Z", "D:00000000000000Z"}:
        return None

    try:
        return fitz.parse_pdf_date(pdf_date_str)
    except Exception:
        try:
            if pdf_date_str.startswith("D:"):
                pdf_date_str = pdf_date_str[2:]
            pdf_date_str = re.sub(r"'(\d{2})'", r"\1", pdf_date_str)
            return date_parser.parse(pdf_date_str, fuzzy=True)
        except Exception as exc:
            logging.debug(f"Failed to parse date '{pdf_date_str}': {exc}")
            return None


def extract_annotations(page, page_number):
    """
    Extract comprehensive annotation information from a page.
    """
    annotations = []
    
    try:
        for annot in page.annots():
            if annot is None:
                continue
                
            try:
                # Basic annotation properties
                annot_info = {
                    'page_number': page_number,
                    'type': annot.type[1] if hasattr(annot, 'type') and annot.type else 'Unknown',
                    'rect': [round(coord, 2) for coord in annot.rect] if hasattr(annot, 'rect') else None,
                    'flags': annot.flags if hasattr(annot, 'flags') else None,
                    
                    # Content and text
                    'contents': annot.contents.strip() if hasattr(annot, 'contents') and annot.contents else None,
                    'text': annot.text if hasattr(annot, 'text') else None,
                    
                    # Appearance properties
                    'colors': {
                        'stroke': annot.colors.get('stroke') if hasattr(annot, 'colors') else None,
                        'fill': annot.colors.get('fill') if hasattr(annot, 'colors') else None
                    },
                    'opacity': annot.opacity if hasattr(annot, 'opacity') else None,
                    'border': annot.border if hasattr(annot, 'border') else None,
                    
                    # Metadata
                    'modification_date': None,  # Will be populated from info
                    'creation_date': None,      # Will be populated from info
                    'popup_rect': [round(coord, 2) for coord in annot.popup_rect] if hasattr(annot, 'popup_rect') else None,
                    'popup': annot.popup if hasattr(annot, 'popup') else None,
                    
                    # Line and vertex properties
                    'vertices': annot.vertices if hasattr(annot, 'vertices') else None,
                    'line_endpoints': annot.line_endpoints if hasattr(annot, 'line_endpoints') else None,
                    
                    # Additional properties
                    'rotation': annot.rotation if hasattr(annot, 'rotation') else None,
                    'quad_points': annot.quad_points if hasattr(annot, 'quad_points') else None,
                    'is_open': annot.is_open if hasattr(annot, 'is_open') else None
                }

                # Extract info dictionary properties
                if hasattr(annot, 'info') and annot.info:
                    info_dict = annot.info
                    info_properties = {
                        'title': info_dict.get('title', ''),
                        'subject': info_dict.get('subject', ''),
                        'creator': info_dict.get('creator', ''),
                        'content': info_dict.get('content', ''),
                        'name': info_dict.get('name', ''),
                        'state': info_dict.get('state', ''),
                        'state_model': info_dict.get('stateModel', '')
                    }
                    annot_info.update(info_properties)

                    # Parse dates
                    creation_date = info_dict.get('creationDate', '')
                    if creation_date:
                        try:
                            parsed_date = parse_pdf_date(creation_date)
                            if parsed_date:
                                annot_info['creation_date'] = parsed_date.strftime('%Y-%m-%d %H:%M:%S')
                        except Exception as e:
                            logging.debug(f"Error parsing creation date: {e}")

                    mod_date = info_dict.get('modDate', '')
                    if mod_date:
                        try:
                            parsed_date = parse_pdf_date(mod_date)
                            if parsed_date:
                                annot_info['modification_date'] = parsed_date.strftime('%Y-%m-%d %H:%M:%S')
                        except Exception as e:
                            logging.debug(f"Error parsing modification date: {e}")

                # Extract colors in hex format
                if annot_info['colors']['stroke']:
                    try:
                        rgb = tuple(int(c * 255) for c in annot_info['colors']['stroke'][:3])
                        annot_info['stroke_color'] = rgb_to_hex(rgb)
                    except Exception as e:
                        logging.debug(f"Error converting stroke color: {e}")

                if annot_info['colors']['fill']:
                    try:
                        rgb = tuple(int(c * 255) for c in annot_info['colors']['fill'][:3])
                        annot_info['fill_color'] = rgb_to_hex(rgb)
                    except Exception as e:
                        logging.debug(f"Error converting fill color: {e}")

                # Extract font properties for FreeText annotations
                if annot_info['type'] == 'FreeText':
                    try:
                        if hasattr(annot, 'info'):
                            da_string = annot.info.get('defaultAppearance', '')
                            if da_string:
                                # Font name and size
                                font_match = re.search(r'/([^\s]+)\s+(\d+)\s+Tf', da_string)
                                if font_match:
                                    annot_info['font_name'] = font_match.group(1)
                                    annot_info['font_size'] = int(font_match.group(2))

                                # Font color
                                color_match = re.search(r'(\d*\.?\d+)\s+(\d*\.?\d+)\s+(\d*\.?\d+)\s+rg', da_string)
                                if color_match:
                                    rgb = tuple(int(float(c) * 255) for c in color_match.groups())
                                    annot_info['font_color'] = rgb_to_hex(rgb)
                    except Exception as e:
                        logging.debug(f"Error extracting font properties: {e}")

                # Clean up the annotation info by removing None and empty values
                cleaned_info = {}
                for key, value in annot_info.items():
                    if value not in (None, '', [], {}, {'stroke': None, 'fill': None}):
                        cleaned_info[key] = value

                if cleaned_info:
                    annotations.append(cleaned_info)
                    logging.debug(f"Successfully extracted annotation from page {page_number}: {cleaned_info['type']}")

            except Exception as e:
                logging.error(f"Error processing annotation on page {page_number}: {e}")
                continue

    except Exception as e:
        logging.error(f"Error accessing annotations on page {page_number}: {e}")
        return []

    return annotations

def extract_pdf_info(pdf_path, max_pages=None):
    logging.info(f"Starting to extract PDF info from: {pdf_path}")
    pdf_data = {'bookmarks': [], 'pages': [], 'annotations': [], 'styled_text': []}

    if not os.path.exists(pdf_path):
        logging.error(f"PDF file not found: {pdf_path}")
        return None

    try:
        with fitz.open(pdf_path) as doc:
            total_pages = len(doc)
            logging.info(f"Successfully opened the PDF. Number of pages: {total_pages}")
        
            # Log PDF metadata for debugging
            logging.info(f"PDF Version: {doc.metadata.get('format', 'Unknown')}")
            logging.info(f"PDF Producer: {doc.metadata.get('producer', 'Unknown')}")

            # Limit the number of pages if max_pages is set
            if max_pages is not None:
                max_pages = min(max_pages, total_pages)
            else:
                max_pages = total_pages

            # Extract bookmarks
            pdf_data['bookmarks'] = doc.get_toc()
            if not pdf_data['bookmarks']:
                logging.info("No bookmarks found in the PDF.")

            all_annotations = []
            all_styled_text = []
            all_pages = []

            # Function to convert color integer to RGB tuple
            def int_to_rgb(color_int):
                """Convert a color integer to an RGB tuple."""
                r = (color_int >> 16) & 255
                g = (color_int >> 8) & 255
                b = color_int & 255
                return (r, g, b)


            # Extract text, annotations, and style attributes for each page
            for page_num in range(max_pages):
                page = doc[page_num]
                page_number = page_num + 1
                logging.info(f"Processing page {page_number}")

                # Extract page text
                page_text = page.get_text().strip()
                page_data = {
                    'page_number': page_number,
                    'text': page_text
                }
                all_pages.append(page_data)

                # Extract annotations using the new function
                page_annotations = extract_annotations(page, page_number)
                all_annotations.extend(page_annotations)

                # Extract styled text
                try:
                    blocks = page.get_text("dict")["blocks"]
                    for block in blocks:
                        if block['type'] == 0:
                            for line in block['lines']:
                                for span in line['spans']:
                                    color = span.get('color', 0)
                                    color_rgb = int_to_rgb(color)
                                    font_color_hex = rgb_to_hex(color_rgb)
                                    span_info = {
                                        'page_number': page_number,
                                        'text': span.get('text', '').strip(),
                                        'font': span.get('font', ''),
                                        'font_size': span.get('size', ''),
                                        'font_color': font_color_hex,
                                        'bbox': [round(coord, 3) for coord in span.get('bbox', [])],
                                    }
                                    # Remove empty fields
                                    span_info = {k: v for k, v in span_info.items() if v not in ['', None]}
                                    all_styled_text.append(span_info)
                except Exception as e:
                    logging.error(f"Error extracting styled text on page {page_number}: {e}", exc_info=True)

            pdf_data['pages'] = all_pages
            pdf_data['annotations'] = all_annotations
            pdf_data['styled_text'] = all_styled_text

            logging.info(f"Extracted {len(all_annotations)} annotations across {max_pages} pages")
            return pdf_data

    except Exception as e:
        logging.error(f"An error occurred while processing the PDF: {e}", exc_info=True)
        return None

def rgb_to_hex(rgb):
    """Convert RGB tuple to hex color code."""
    try:
        return '#{:02X}{:02X}{:02X}'.format(*rgb)
    except Exception:
        return ''

def clean_value_for_excel(value):
    """Clean value for Excel export."""
    if value is None:
        return ''
    if isinstance(value, (list, dict)):
        return str(value).replace('\x00', '').replace('\r', '')
    return str(value).replace('\x00', '').replace('\r', '')



def save_to_excel(formatted_data, output_path):
    """Save formatted JSON data to Excel with enhanced styling and filters."""
    try:
        # Create workbook in normal mode for styling support
        wb = Workbook()
        
        # Define common styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_alignment = Alignment(vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def apply_sheet_styling(ws, start_col, end_col):
            """Helper function to apply consistent styling to sheets"""
            # Apply autofilter
            ws.auto_filter.ref = ws.dimensions
            
            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Format data cells
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = data_alignment
                    cell.border = border
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = list(col)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap width at 50
                ws.column_dimensions[col[0].column_letter].width = adjusted_width
            
            # Freeze top row
            ws.freeze_panes = 'A2'
            
            # Set zoom level
            ws.sheet_view.zoomScale = 85

        # Remove default sheet
        wb.remove(wb.active)

        # Create Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Component", "Count"])
        for component in formatted_data['summary']['components']:
            ws_summary.append([
                str(component.get('Component', '')),
                str(component.get('Count', ''))
            ])
        apply_sheet_styling(ws_summary, 'A', 'B')

        # Create Annotations sheet
        if formatted_data['sheets']['annotations']:
            ws_annotations = wb.create_sheet("Annotations")
            
            headers = [
                'Page Number', 'Annotation Type', 'Content', 'Position',
                'flags', 'colors stroke', 'colors fill', 'Stroke Color',
                'Opacity', 'Border Width', 'Border Dashes', 'Border Style',
                'Border Clouds', 'Rotation', 'Flags', 'Is Open', 'Popup Rectangle'
            ]
            
            ws_annotations.append(headers)
            
            for annot in formatted_data['sheets']['annotations']:
                row = []
                for header in headers:
                    value = annot.get(header, '')
                    if isinstance(value, (list, dict)):
                        value = str(value)
                    row.append(str(value))
                ws_annotations.append(row)
            
            apply_sheet_styling(ws_annotations, 'A', 'Q')

        # Create Bookmarks sheet
        if formatted_data['sheets']['bookmarks']:
            ws_bookmarks = wb.create_sheet("Bookmarks")
            ws_bookmarks.append(['Level', 'Title', 'Page'])
            for bookmark in formatted_data['sheets']['bookmarks']:
                ws_bookmarks.append([
                    str(bookmark.get('Level', '')),
                    str(bookmark.get('Title', '')),
                    str(bookmark.get('Page', ''))
                ])
            apply_sheet_styling(ws_bookmarks, 'A', 'C')

        # Create Pages sheet
        if formatted_data['sheets']['pages']:
            ws_pages = wb.create_sheet("Pages")
            ws_pages.append(['Page Number', 'Text'])
            for page in formatted_data['sheets']['pages']:
                ws_pages.append([
                    str(page.get('Page Number', '')),
                    str(page.get('Text', '')).replace('\x00', '').replace('\r', '')
                ])
            apply_sheet_styling(ws_pages, 'A', 'B')

        # Create Styled Text sheet
        if formatted_data['sheets']['styled_text']:
            ws_styled = wb.create_sheet("Styled Text")
            ws_styled.append(['Page Number', 'Text', 'Font', 'Font Size', 'Font Color', 'Position'])
            for text in formatted_data['sheets']['styled_text']:
                ws_styled.append([
                    str(text.get('Page Number', '')),
                    str(text.get('Text', '')),
                    str(text.get('Font', '')),
                    str(text.get('Font Size', '')),
                    str(text.get('Font Color', '')),
                    str(text.get('Position', '')).replace('=', '')  # Prevent formula injection
                ])
            apply_sheet_styling(ws_styled, 'A', 'F')

        # Save workbook
        wb.save(filename=output_path)
        logging.info(f"Successfully saved Excel file to: {output_path}")
        
    except Exception as e:
        logging.error(f"Failed to create Excel workbook: {e}", exc_info=True)
        raise

def clean_data_for_excel(data):
    """Clean data before converting to DataFrame."""
    if isinstance(data, dict):
        return {k: clean_data_for_excel(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [clean_data_for_excel(item) for item in data]
    elif isinstance(data, (int, float)):
        return str(data)
    elif isinstance(data, str):
        # Remove problematic characters
        return data.replace('\x00', '').replace('\r', '')
    elif data is None:
        return ''
    else:
        return str(data)

def process_pdf(
    pdf_path, 
    output_dir=None, 
    max_pages=None,
    formatted_json_name=None,
    excel_name=None
):
    """
    Main function implementing the three-step process.
    
    Args:
        pdf_path (str): Path to the input PDF file
        output_dir (str, optional): Directory for output files. If None, uses PDF directory
        max_pages (int, optional): Maximum number of pages to process
        formatted_json_name (str, optional): Name for the formatted JSON file
        excel_name (str, optional): Name for the Excel output file
    """
    try:
        # Generate file paths
        if output_dir is None:
            output_dir = os.path.dirname(pdf_path)
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        pdf_filename = os.path.basename(pdf_path)
        base_name = os.path.splitext(pdf_filename)[0]
        
        # Generate file paths
        json_path = os.path.join(output_dir, f"{base_name}_raw.json")
        formatted_json_path = os.path.join(
            output_dir,
            formatted_json_name if formatted_json_name else f"{base_name}_tabular.json"
        )
        excel_path = os.path.join(
            output_dir,
            excel_name if excel_name else f"{base_name}_report.xlsx"
        )

        # Step 1: Extract PDF info and save to JSON
        logging.info("Step 1: Extracting PDF info...")
        result = extract_pdf_info(pdf_path, max_pages)
        if result:
            save_to_json(result, json_path)
            logging.info(f"Step 1 complete: Raw data saved to {json_path}")

            # Step 2: Create formatted JSON
            logging.info("Step 2: Creating formatted JSON...")
            formatted_data = create_formatted_json(json_path, formatted_json_path)
            logging.info(f"Step 2 complete: Formatted data saved to {formatted_json_path}")

            # Clean data before saving to Excel
            cleaned_data = clean_data_for_excel(formatted_data)

            # Step 3: Generate Excel file
            logging.info("Step 3: Generating Excel file...")
            save_to_excel(cleaned_data, excel_path)
            logging.info(f"Step 3 complete: Excel file saved to {excel_path}")
        else:
            logging.error("Failed to extract PDF information")

    except Exception as e:
        logging.error(f"Error in process_pdf: {e}", exc_info=True)

    logging.info("Script finished")

def save_to_json(data, output_path):
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            logging.info(f"Extracted information saved to JSON: {output_path}")
    except Exception as e:
        logging.error(f"Failed to save JSON file: {e}", exc_info=True)

def create_formatted_json(input_json_path, output_json_path):
    """
    Step 2: Read acrf.json and create p_01_02_acrf_excel_ready_format.json with Excel-ready format
    """
    try:
        # Read the original JSON file
        with open(input_json_path, 'r', encoding='utf-8') as f:
            raw_data = json.load(f)
        
        # Initialize Excel-ready format
        excel_ready_data = {
            'summary': {
                'components': [
                    {'Component': 'Total Pages', 'Count': len(raw_data.get('pages', []))},
                    {'Component': 'Annotations', 'Count': len(raw_data.get('annotations', []))},
                    {'Component': 'Bookmarks', 'Count': len(raw_data.get('bookmarks', []))},
                    {'Component': 'Styled Text Elements', 'Count': len(raw_data.get('styled_text', []))}
                ]
            },
            'sheets': {
                'annotations': [],
                'bookmarks': [],
                'pages': [],
                'styled_text': []
            }
        }

        # Format annotations
        for annot in raw_data.get('annotations', []):
            formatted_annot = {
                'Page Number': annot.get('page_number', ''),
                'Annotation Type': annot.get('type', ''),
                'Content': annot.get('content', ''),
                'Position': clean_value_for_excel(annot.get('rect', '')),
                'flags': annot.get('flags', ''),
                'colors stroke': clean_value_for_excel(annot.get('colors', {}).get('stroke', '')),
                'colors fill': clean_value_for_excel(annot.get('colors', {}).get('fill', '')),
                'Stroke Color': annot.get('stroke_color', ''),
                'Opacity': annot.get('opacity', ''),
                'Border Width': annot.get('border', {}).get('width', ''),
                'Border Dashes': clean_value_for_excel(annot.get('border', {}).get('dashes', '')),
                'Border Style': annot.get('border', {}).get('style', ''),
                'Border Clouds': annot.get('border', {}).get('clouds', ''),
                'Rotation': annot.get('rotation', ''),
                'Flags': annot.get('flags', ''),
                'Is Open': annot.get('is_open', ''),
                'Popup Rectangle': clean_value_for_excel(annot.get('popup_rect', ''))
            }
            
            # Add to sheets
            excel_ready_data['sheets']['annotations'].append(formatted_annot)

        # Format bookmarks
        for bookmark in raw_data.get('bookmarks', []):
            if isinstance(bookmark, (list, tuple)) and len(bookmark) >= 3:
                formatted_bookmark = {
                    'Level': bookmark[0],
                    'Title': clean_value_for_excel(bookmark[1]),
                    'Page': bookmark[2]
                }
                excel_ready_data['sheets']['bookmarks'].append(formatted_bookmark)

        # Format pages
        for page in raw_data.get('pages', []):
            formatted_page = {
                'Page Number': page.get('page_number', ''),
                'Text': clean_value_for_excel(page.get('text', ''))
            }
            if formatted_page['Text']:  # Only add if there's text content
                excel_ready_data['sheets']['pages'].append(formatted_page)

        # Format styled text
        for text in raw_data.get('styled_text', []):
            formatted_text = {
                'Page Number': text.get('page_number', ''),
                'Text': clean_value_for_excel(text.get('text', '')),
                'Font': text.get('font', ''),
                'Font Size': text.get('font_size', ''),
                'Font Color': text.get('font_color', ''),
                'Position': str(text.get('bbox', ''))
            }
            excel_ready_data['sheets']['styled_text'].append(formatted_text)

        # Save the formatted JSON
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(excel_ready_data, f, ensure_ascii=False, indent=2)
        
        logging.info(f"Successfully created formatted JSON at: {output_json_path}")
        return excel_ready_data

    except Exception as e:
        logging.error(f"Error creating formatted JSON: {e}", exc_info=True)
        raise

if __name__ == "__main__":
    # Determine repository root relative to this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.abspath(os.path.join(script_dir, os.pardir))

    # Use paths relative to the repository
    pdf_path = os.path.join(repo_root, "data", "acrf.pdf")
    output_dir = os.path.join(repo_root, "output")

    process_pdf(
        pdf_path=pdf_path,
        output_dir=output_dir,
    )
